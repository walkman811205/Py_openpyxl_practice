# -*- coding: utf-8 -*-
import openpyxl
import datetime
import os

DISK='D:/'
d=-1
day = (datetime.datetime.now()+datetime.timedelta(days=d)).strftime("%Y%m%d")
path = DISK+day

stock = '/現股-'+day+'.xlsx'
warrant = '/權證-'+day+'.xlsx'
sm_buy = '/標的物(購)-'+day+'.xlsx'
sm_sale = '/標的物(售)-'+day+'.xlsx'
sm_cow = '/標的物(牛)-'+day+'.xlsx'
sm_bear = '/標的物(熊)-'+day+'.xlsx'
sm_top20_buy = '/標的物-Top20(購)-'+day+'.xlsx'
sm_top20_sale = '/標的物-Top20(售)-'+day+'.xlsx'

subject_title=['股票代號','股票名稱','證券商','累積數量','交易數量','最後揭示買量','最後揭示賣量','外資自營商買進股數(不含外資自營商)','外資自營商賣出股數(不含外資自營商)','外資自營商買賣超股數(不含外資自營商)','自營商買賣超股數','自營商買進股數(避險)','自營商賣出股數(避險)','自營商買賣超股數(避險)','三大法人買賣超股數']
subject20_title=['股票代號','股票名稱','買進/賣出股數','買賣超股數','漲跌天數','1日漲幅','5日漲幅','自營商(避險)比例(%)','','股票代號','股票名稱','序號','買點','價格','買進股數','賣出股數']

if not os.path.isdir(path):
    os.makedirs(path)
    print('----建立成功----')
else:
    print('目錄已建立過位於-'+path)

def create_execel(day,DISK):
    file = openpyxl.Workbook()

    sheet = file.active
    sheet.title = '現股'
    file.save(path+stock)
    sheet.title = '權證'
    file.save(path+warrant)
    sheet.title = '標的物(購)'
    file.save(path+sm_buy)
    sheet.title = '標的物(售)'
    file.save(path+sm_sale)
    sheet.title = '標的物(牛)'
    file.save(path+sm_cow)
    sheet.title = '標的物(熊)'
    file.save(path+sm_bear)
    sheet.title = '標的物-Top20(購)'
    file.save(path+sm_top20_buy)
    sheet.title = '標的物-Top20(售)'
    file.save(path+sm_top20_sale)

def set_execel(day,DISK):
    file=openpyxl.load_workbook(path+'/現股-'+day+'.xlsx')
    sheet = file['現股']
    sheet.append(['證券代號','證券名稱','外陸資買進股數','外陸資賣出股數','外陸資買賣超股數','投信買進股數','投信賣出股數','投信買賣超股數','自營商買賣超股數','自營商買進股數(自行買賣)','自營商賣出股數(自行買賣)','自營商買賣超股數(自行買賣)','自營商買進股數(避險)','自營商賣出股數(避險)','自營商買賣超股數(避險)','三大法人買賣超股數','自營商避險比例(%)','5天總計(排除避險)','5日中位數','5日法人占比量(%)','10天總計(排除避險)','10日中位數','10日法人占比量(%)','20天總計(排除避險)','20日中位數','收盤價','上布林斜率','上布林','K棒','狀態','日期'])
    file.save(path+stock)

    file=openpyxl.load_workbook(path+'/權證-'+day+'.xlsx')
    sheet = file['權證']
    sheet.append(['證券代號','證券名稱','外資自營商買進股數(不含外資自營商)','外資自營商賣出股數(不含外資自營商)','外資自營商買賣超股數(不含外資自營商)','自營商買賣超股數','自營商買進股數(避險)','自營商賣出股數(避險)','自營商買賣超股數(避險)','三大法人買賣超股數','日期'])
    file.save(path+warrant)

    file=openpyxl.load_workbook(path+'/標的物(購)-'+day+'.xlsx')
    sheet = file['標的物(購)']
    sheet.append(subject_title)
    file.save(path+sm_buy)

    file=openpyxl.load_workbook(path+'/標的物(售)-'+day+'.xlsx')
    sheet = file['標的物(售)']
    sheet.append(subject_title)
    file.save(path+sm_sale)

    file=openpyxl.load_workbook(path+'/標的物(牛)-'+day+'.xlsx')
    sheet = file['標的物(牛)']
    sheet.append(subject_title)
    file.save(path+sm_cow)

    file=openpyxl.load_workbook(path+'/標的物(熊)-'+day+'.xlsx')
    sheet = file['標的物(熊)']
    sheet.append(subject_title)
    file.save(path+sm_bear)

    file=openpyxl.load_workbook(path+'/標的物-Top20(購)-'+day+'.xlsx')
    sheet = file['標的物-Top20(購)']
    sheet.append(subject20_title)
    file.save(path+sm_top20_buy)

    file=openpyxl.load_workbook(path+'/標的物-Top20(售)-'+day+'.xlsx')
    sheet = file['標的物-Top20(售)']
    sheet.append(subject20_title)
    file.save(path+sm_top20_sale)


create_execel(day,DISK)
set_execel(day,DISK)