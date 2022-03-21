# -*- coding: utf-8 -*-
import openpyxl
import datetime
import pymysql
import dbconfig


DISK='D:/'
d=-1
day = (datetime.datetime.now()+datetime.timedelta(days=d)).strftime("%Y%m%d")
path = DISK+day

dbconnect = dbconfig.setting
conn = pymysql.connect(**dbconnect)
cursor = conn.cursor()

stockName_table={'SP黃豆':'街口S&P黃豆','元石油':'元大S&P石油','台灣50':'元大台灣50','大銀':'大銀微系統','道瓊反':'國泰美國道瓊反1','寶滬深':'元大滬深300反1','元黃金':'元大S&P黃金','FANG+':'統一FANG+','上証反':'富邦上証反1','金融類':'元大MSCI金融','S&P500':'元大S&P500','滬深2X':'元大滬深300正2','復盛':'復盛應用','京元電':'京元電子','臺指反':'國泰臺灣加權反1','上海銀':'上海商銀','T50正2':'元大台灣50正2','上証2X':'富邦上証正2','寶齡':'寶齡富錦','道瓊銀':'元大道瓊白銀','T50反1':'元大台灣50反1','上緯控':'上緯投控','道瓊銅':'街口道瓊銅','FBVIX':'富邦VIX','高鐵':'台灣高鐵','美指2X':'元大美元指數正2','長華':'長華*','SGBR2X':'街口布蘭特油正2','昇陽半':'昇陽半導體','瑞祺':'瑞祺電通','S&P正2':'元大S&P500正2','深中小':'群益深証中小','元金2X':'元大S&P黃金正2','元大滬深300反1':'滬深反','香港2X':'FH香港正2','A50':'國泰中國A50','A50反1':'國泰中國A50反1','元上證':'元大上證50','高股息':'元大高股息','東鋼':'東和鋼鐵','日月光':'日月光投控','A50正2':'國泰中國A50正2','S&P反1':'元大S&P500反1','FH中5G':'FH中國5G','康全電':'康全電訊','藏壽司':'亞洲藏壽司','雍智':'雍智科技','昇佳電':'昇佳電子','元油反':'元大S&P原油反1','FB上証':'富邦上証','中橡':'國際中橡','和潤':'和潤企業','晨訊科':'晨訊科-DR','恒指':'FH香港','電子類':'元大電子'}

def change(data):
    y=''
    for i in data:
        if i!=',':
            y=y+i
    return float(y)

def stock_code(NAME):
    name=stockName_table.get(NAME)
    if name==None:
        name=NAME
    cursor.execute("select security_code from {} where security_name='{}'".format(table[0],name))
    if cursor.fetchone()!= None:
        return cursor.fetchone()[0],name
    else:
        name1=str(name)+'-KY'
        cursor.execute("select security_code from {} where security_name='{}'".format(table[0],name1))
        if cursor.fetchone()!= None:
            return cursor.fetchone()[0],name1
        else:
            #print(name)
            return [' ',name]

def mark_data():
    file=openpyxl.load_workbook(path+'/標的物(購)-'+day+'.xlsx')
    file1=openpyxl.load_workbook(path+'/標的物(售)-'+day+'.xlsx')
    file2=openpyxl.load_workbook(path+'/標的物(牛)-'+day+'.xlsx')
    file3=openpyxl.load_workbook(path+'/標的物(熊)-'+day+'.xlsx')
    greenfill=openpyxl.styles.PatternFill(fill_type='solid',fgColor='00BB00')
    redfill=openpyxl.styles.PatternFill(fill_type='solid',fgColor='FF9797')
    
    sheet0 = file['標的物(購)'] 
    sheet1 = file1['標的物(售)'] 
    sheet2 = file2['標的物(牛)'] 
    sheet3 = file3['標的物(熊)'] 
    k0=sheet0.iter_cols(min_row=2,max_row=sheet0.max_row,min_col=2,max_col=2)
    k1=sheet1.iter_cols(min_row=2,max_row=sheet1.max_row,min_col=2,max_col=2)
    k2=sheet2.iter_cols(min_row=2,max_row=sheet2.max_row,min_col=2,max_col=2)
    k3=sheet3.iter_cols(min_row=2,max_row=sheet3.max_row,min_col=2,max_col=2)
    for i in k0:
        for j in i:
            if j.value=='紅K':
                result_excel=sheet0.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet0.min_column,max_col=sheet0.max_column)
                for x in result_excel:
                    for y in x:
                        y.fill=redfill
            elif j.value=='黑K':
                result_excel=sheet0.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet0.min_column,max_col=sheet0.max_column)
                for x in result_excel:
                    for y in x:
                        y.fill=greenfill
    for i in k1:
        for j in i:
            if j.value=='紅K':
                result_excel=sheet1.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet1.min_column,max_col=sheet1.max_column)
                for x in result_excel:
                    for y in x:
                        y.fill=redfill
            elif j.value=='黑K':
                result_excel=sheet1.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet1.min_column,max_col=sheet1.max_column)
                for x in result_excel:
                    for y in x:
                        y.fill=greenfill
    for i in k2:
        for j in i:
            if j.value=='紅K':
                result_excel=sheet2.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet2.min_column,max_col=sheet2.max_column)
                for x in result_excel:
                    for y in x:
                        y.fill=redfill
            elif j.value=='黑K':
                result_excel=sheet2.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet2.min_column,max_col=sheet2.max_column)
                for x in result_excel:
                    for y in x:
                        y.fill=greenfill
    for i in k3:
        for j in i:
            if j.value=='紅K':
                result_excel=sheet3.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet3.min_column,max_col=sheet3.max_column)
                for x in result_excel:
                    for y in x:
                        y.fill=redfill
            elif j.value=='黑K':
                result_excel=sheet3.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet3.min_column,max_col=sheet3.max_column)
                for x in result_excel:
                    for y in x:
                        y.fill=greenfill
    file.save(path+'/標的物(購)-'+day+'.xlsx')    
    file1.save(path+'/標的物(售)-'+day+'.xlsx')
    file2.save(path+'/標的物(牛)-'+day+'.xlsx')
    file3.save(path+'/標的物(熊)-'+day+'.xlsx')

file=openpyxl.load_workbook(path+'/標的物(購)-'+day+'.xlsx')
file1=openpyxl.load_workbook(path+'/標的物(售)-'+day+'.xlsx')
file2=openpyxl.load_workbook(path+'/標的物(牛)-'+day+'.xlsx')
file3=openpyxl.load_workbook(path+'/標的物(熊)-'+day+'.xlsx')
sheet0 = file['標的物(購)'] 
sheet1 = file1['標的物(售)'] 
sheet2 = file2['標的物(牛)'] 
sheet3 = file3['標的物(熊)'] 

for table in[['daily_quotes','trading_detail'],['daily_otc','trading_otc']]:

    #cursor.execute("select security_code,security_name,trade_value,lbb_volume,lba_volume from {} where length(security_code)=6 and quotes_date='{}'".format(table,day))
    mysql=f"select {table[0]}.security_name,{table[0]}.trade_value,{table[0]}.lbb_volume,{table[0]}.lba_volume,{table[1]}.fd_totalbuy,{table[1]}.fd_totalsell,{table[1]}.fd_difference,{table[1]}.d_difference,{table[1]}.dh_totalbuy,{table[1]}.dh_totalsell,{table[1]}.dh_difference,{table[1]}.total_difference from {table[0]} inner join {table[1]} on {table[0]}.security_code={table[1]}.security_code and length({table[0]}.security_code)=6 and {table[0]}.quotes_date='{day}' and {table[1]}.trade_date='{day}';"
    cursor.execute(mysql)
#     result_data={}
    warrant_buy=[]
    warrant_sell=[]
    warrant_cow=[]
    warrant_bear=[]
#     all_data=[]
    for i in cursor.fetchall():
        if '購' in i[0]:
            warrant_buy.append(list(i))
        elif '售' in i[0]:
            warrant_sell.append(list(i))
        elif '牛' in i[0]:
            warrant_cow.append(list(i))
        elif '熊' in i[0]:
            warrant_bear.append(list(i))
            
    for warrant_type,sheet_num in zip([warrant_buy,warrant_sell,warrant_cow,warrant_bear],range(4)):
        subject=[]
        issue=[]
        for i in warrant_type:
            for j in ['購','售','牛','熊']:
                if j in i[0]:
                    temp=i[0].split(j)
                    warrant_issue=temp[0][-4:-2]
                    subject_name=temp[0][0:-4]
                    subject.append(subject_name)
                    issue.append(warrant_issue)
        data_issue=list(set(issue))
        data_stock=list(set(subject))
        
        result=[]
        n=0
        for x in data_stock:
            n+=1
            for b in range(12):
                #number,trade_value,lbb_volume,lba_volume,fd_totalbuy,fd_totalsell,fd_difference,d_difference,dh_totalbuy,dh_totalsell,dh_difference,total_difference
                b=b+1    
                locals()['total_add'+str(b)]=0
            for y in data_issue:
                temp_data=[]
                temp=x+y
                for i in warrant_type:
                    if temp in i[0]:
                        temp_data.append(i)
                for a in range(12):
                    a=a+1    
                    locals()['add'+str(a)]=0
                if len(temp_data)>0:
                    for j in temp_data:
                        #trade_value,lbb_volume,lba_volume,fd_totalbuy,fd_totalsell,fd_difference,d_difference,dh_totalbuy,dh_totalsell,dh_difference,total_difference
                        for a in range(11):
                            a=a+1    
                            locals()['add'+str(a)]=locals()['add'+str(a)]+change(j[a])
                            locals()['total_add'+str(a+1)]=locals()['total_add'+str(a+1)]+change(j[a])
                    locals()['total_add1']=locals()['total_add1']+int(len(temp_data))
                    result.append([stock_code(x)[0],stock_code(x)[1],y,len(temp_data),locals()['add1'],format(locals()['add2'],','),format(locals()['add3'],','),format(locals()['add4'],','),format(locals()['add5'],','),format(locals()['add6'],','),format(locals()['add7'],','),format(locals()['add8'],','),format(locals()['add9'],','),format(locals()['add10'],','),format(locals()['add11'],',')])
#             print(result)
            cursor.execute("select closing_price,opening_price from {} where security_code='{}' and quotes_date='{}'".format(table[0],stock_code(x)[0],day))
            k_data=cursor.fetchone()
            if k_data!=None:
                if change(k_data[0])>=change(k_data[1]):
                    k_state='紅K'
                else:
                    k_state='黑K'
                result.append([k_data[0],k_state,'總計',locals()['total_add1'],format(locals()['total_add2'],','),format(locals()['total_add3'],','),format(locals()['total_add4'],','),format(locals()['total_add5'],','),format(locals()['total_add6'],','),format(locals()['total_add7'],','),format(locals()['total_add8'],','),format(locals()['total_add9'],','),format(locals()['total_add10'],','),format(locals()['total_add11'],','),format(locals()['total_add12'],','),stock_code(x)[0]])
            else:
                result.append(['','','總計',locals()['total_add1'],format(locals()['total_add2'],','),format(locals()['total_add3'],','),format(locals()['total_add4'],','),format(locals()['total_add5'],','),format(locals()['total_add6'],','),format(locals()['total_add7'],','),format(locals()['total_add8'],','),format(locals()['total_add9'],','),format(locals()['total_add10'],','),format(locals()['total_add11'],','),format(locals()['total_add12'],','),stock_code(x)[0]])   
        for i in result:
            locals()['sheet'+str(sheet_num)].append(i)
            # print(n)
            n+=1

file.save(path+'/標的物(購)-'+day+'.xlsx')    
file1.save(path+'/標的物(售)-'+day+'.xlsx')
file2.save(path+'/標的物(牛)-'+day+'.xlsx')
file3.save(path+'/標的物(熊)-'+day+'.xlsx')

mark_data()
cursor.close()
conn.close()