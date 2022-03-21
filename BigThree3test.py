# -*- coding: utf-8 -*-
"""
Created on Mon Dec 28 11:33:39 2020

@author: Alice
"""
import datetime
import pymysql
import openpyxl
import dbconfig

dbconnect = dbconfig.setting

#day = (datetime.datetime.now()+datetime.timedelta(days=d)).strftime("%Y-%m-%d %H:%M:%S")
#day = (datetime.datetime.now()+datetime.timedelta(days=d)).strftime("%Y-%m-%d")
#print(day)

d=-1
day = (datetime.datetime.now()+datetime.timedelta(days=d)).strftime("%Y-%m-%d")


#day=setTime0.strftime('%Y%m%d')


DISK='D:/'

def main(day,DISK):
    d=-4
    day = (datetime.datetime.now()+datetime.timedelta(days=d)).strftime("%Y%m%d")
    print(day)

    DISK='D:/'
    print("Bigthree3")

    conn=pymysql.connect(**dbconnect)
    cursor=conn.cursor()
    stockName_table={'SP黃豆':'街口S&P黃豆','元石油':'元大S&P石油','台灣50':'元大台灣50','大銀':'大銀微系統','道瓊反':'國泰美國道瓊反1','寶滬深':'元大滬深300反1','元黃金':'元大S&P黃金','FANG+':'統一FANG+','上証反':'富邦上証反1','金融類':'元大MSCI金融','S&P500':'元大S&P500','滬深2X':'元大滬深300正2','復盛':'復盛應用','京元電':'京元電子','臺指反':'國泰臺灣加權反1','上海銀':'上海商銀','T50正2':'元大台灣50正2','上証2X':'富邦上証正2','寶齡':'寶齡富錦','道瓊銀':'元大道瓊白銀','T50反1':'元大台灣50反1','上緯控':'上緯投控','道瓊銅':'街口道瓊銅','FBVIX':'富邦VIX','高鐵':'台灣高鐵','美指2X':'元大美元指數正2','長華':'長華*','SGBR2X':'街口布蘭特油正2','昇陽半':'昇陽半導體','瑞祺':'瑞祺電通','S&P正2':'元大S&P500正2','深中小':'群益深証中小','元金2X':'元大S&P黃金正2','元大滬深300反1':'滬深反','香港2X':'FH香港正2','A50':'國泰中國A50','A50反1':'國泰中國A50反1','元上證':'元大上證50','高股息':'元大高股息','東鋼':'東和鋼鐵','日月光':'日月光投控','A50正2':'國泰中國A50正2','S&P反1':'元大S&P500反1','FH中5G':'FH中國5G','康全電':'康全電訊','藏壽司':'亞洲藏壽司','雍智':'雍智科技','昇佳電':'昇佳電子','元油反':'元大S&P原油反1','FB上証':'富邦上証','中橡':'國際中橡','和潤':'和潤企業','晨訊科':'晨訊科-DR','恒指':'FH香港','電子類':'元大電子'}
    subject_title=['股票代號','股票名稱','證券商','累積數量','交易數量','最後揭示買量','最後揭示賣量','外資自營商買進股數(不含外資自營商)','外資自營商賣出股數(不含外資自營商)','外資自營商買賣超股數(不含外資自營商)','自營商買賣超股數','自營商買進股數(避險)','自營商賣出股數(避險)','自營商買賣超股數(避險)','三大法人買賣超股數']
    
    # 建立一個新的execel工作表
    file=openpyxl.Workbook()
    sheet=file.active
    sheet.title='現股'
    file.create_sheet(title='權證',index=1)
    file.create_sheet(title='標的物(購)',index=2)
    file.create_sheet(title='標的物(售)',index=3)
    file.create_sheet(title='標的物(牛)',index=4)
    file.create_sheet(title='標的物(熊)',index=5)
    file.create_sheet(title='標的物-Top20(購)',index=6)
    file.create_sheet(title='標的物-Top20(售)',index=7)
    file.save(DISK+day+'/三大法人買賣超日報3.0-'+day+'.xlsx')
    
    #打開execel寫入分頁
    file=openpyxl.load_workbook(DISK+day+'/三大法人買賣超日報3.0-'+day+'.xlsx')
    sheet = file['現股']
    sheet1 = file['權證']
    sheet2 = file['標的物(購)']
    sheet3 = file['標的物(售)']
    sheet4 = file['標的物(牛)']
    sheet5 = file['標的物(熊)']

    sheet.append(['證券代號','證券名稱','外陸資買進股數','外陸資賣出股數','外陸資買賣超股數','投信買進股數','投信賣出股數','投信買賣超股數','自營商買賣超股數','自營商買進股數(自行買賣)','自營商賣出股數(自行買賣)','自營商買賣超股數(自行買賣)','自營商買進股數(避險)','自營商賣出股數(避險)','自營商買賣超股數(避險)','三大法人買賣超股數','自營商避險比例(%)','5天總計(排除避險)','5日中位數','5日法人占比量(%)','10天總計(排除避險)','10日中位數','10日法人占比量(%)','20天總計(排除避險)','20日中位數','收盤價','上布林斜率','上布林','K棒','狀態','日期'])
    sheet1.append(['證券代號','證券名稱','外資自營商買進股數(不含外資自營商)','外資自營商賣出股數(不含外資自營商)','外資自營商買賣超股數(不含外資自營商)','自營商買賣超股數','自營商買進股數(避險)','自營商賣出股數(避險)','自營商買賣超股數(避險)','三大法人買賣超股數','日期'])
    sheet2.append(subject_title)
    sheet3.append(subject_title)
    sheet4.append(subject_title)
    sheet5.append(subject_title)
    
    def stock_code(NAME):
        name=stockName_table.get(NAME)
        if name==None:
            name=NAME
        cursor.execute("select security_code from {} where security_name='{}'".format(table[0],name))
        if cursor.fetchone()!= None:
            return cursor.fetchone()[0],name
        else:
            name1=str(name)+'-KY'
            cursor.execute("select security_code from {} where security_name='{}'".format(table[0],name1))
            if cursor.fetchone()!= None:
                return cursor.fetchone()[0],name1
            else:
                #print(name)
                return [' ',name]

    def stock_name(code):
        cursor.execute("select security_name from trading_detail where security_code='{}'".format(code))
        name=cursor.fetchone()
        if name==None:
            cursor.execute("select security_name from trading_otc where security_code='{}'".format(code))
            name=cursor.fetchone()
            if name==None:
                return ''
        return name[0]

    def change(data):
        y=''
        for i in data:
            if i!=',':
                y=y+i
        return float(y)

    def change_int(data):
        y=''
        for i in data:
            if i!=',':
                y=y+i
        return int(float(y))

    def increase(code,n):
        s=0
        for i in range(n):
            incre=0
            cursor.execute("select dir,change_ from daily_quotes where security_code='{}' and quotes_date='{}'".format(code, all_day[-1-i]))
            temp=cursor.fetchone()
            if temp==None:
                cursor.execute("select dir from daily_otc where security_code='{}' and quotes_date='{}'".format(code, all_day[-1-i]))
                temp=cursor.fetchone()
                if temp==None:
                    return ''
                else:
#                    incre=float(temp[0])
                    incre=(temp[0])
            else:
                if temp[0]=='+':
                    incre=float(temp[1])
                elif temp[0]=='-':
                    incre=-float(temp[1])
                else:
                    incre=float(temp[1])
                s=s+incre
        return round(s/n,2)
    
    def k(code):
        num=0
        cursor.execute("select opening_price,closing_price from daily_quotes where security_code='{}' and quotes_date='{}'".format(code, day))
        temp=cursor.fetchone()
        if temp!=None:
            t='daily_quotes'
        else:
            t='daily_otc'
        for d in all_day[::-1]:
            cursor.execute("select opening_price,closing_price from {} where security_code='{}' and quotes_date='{}'".format(t,code, d))
            temp=cursor.fetchone()
            if change(temp[0])<=change(temp[1]):
                num+=1
            else:
                if num==0:
                    num-=1
                    continue
                else:
                    break
        return num
    
    def excel():
        file=openpyxl.load_workbook(DISK+day+'/三大法人買賣超日報3.0-'+day+'.xlsx')
        greenfill=openpyxl.styles.PatternFill(fill_type='solid',fgColor='00BB00')
        redfill=openpyxl.styles.PatternFill(fill_type='solid',fgColor='FF9797')
        
        sheet=file['現股']
        sheet1=file['標的物(購)']
        sheet2=file['標的物(售)']
        sheet3=file['標的物(牛)']
        sheet4=file['標的物(熊)']
        k=sheet.iter_cols(min_row=2,max_row=sheet.max_row,min_col=29,max_col=29)
        k1=sheet1.iter_cols(min_row=2,max_row=sheet1.max_row,min_col=2,max_col=2)
        k2=sheet2.iter_cols(min_row=2,max_row=sheet2.max_row,min_col=2,max_col=2)
        k3=sheet3.iter_cols(min_row=2,max_row=sheet3.max_row,min_col=2,max_col=2)
        k4=sheet4.iter_cols(min_row=2,max_row=sheet4.max_row,min_col=2,max_col=2)
        for i in k:
            for j in i:
                if j.value=='紅K':
                    result_excel=sheet.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet.min_column,max_col=sheet.max_column)
                    for x in result_excel:
                        for y in x:
                            y.fill=redfill
                else:
                    result_excel=sheet.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet.min_column,max_col=sheet.max_column)
                    for x in result_excel:
                        for y in x:
                            y.fill=greenfill
        for i in k1:
            for j in i:
                if j.value=='紅K':
                    result_excel=sheet1.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet1.min_column,max_col=sheet1.max_column)
                    for x in result_excel:
                        for y in x:
                            y.fill=redfill
                elif j.value=='黑K':
                    result_excel=sheet1.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet1.min_column,max_col=sheet1.max_column)
                    for x in result_excel:
                        for y in x:
                            y.fill=greenfill
        for i in k2:
            for j in i:
                if j.value=='紅K':
                    result_excel=sheet2.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet2.min_column,max_col=sheet2.max_column)
                    for x in result_excel:
                        for y in x:
                            y.fill=redfill
                elif j.value=='黑K':
                    result_excel=sheet2.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet2.min_column,max_col=sheet2.max_column)
                    for x in result_excel:
                        for y in x:
                            y.fill=greenfill
        for i in k3:
            for j in i:
                if j.value=='紅K':
                    result_excel=sheet3.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet3.min_column,max_col=sheet3.max_column)
                    for x in result_excel:
                        for y in x:
                            y.fill=redfill
                elif j.value=='黑K':
                    result_excel=sheet3.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet3.min_column,max_col=sheet3.max_column)
                    for x in result_excel:
                        for y in x:
                            y.fill=greenfill
        for i in k4:
            for j in i:
                if j.value=='紅K':
                    result_excel=sheet4.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet4.min_column,max_col=sheet4.max_column)
                    for x in result_excel:
                        for y in x:
                            y.fill=redfill
                elif j.value=='黑K':
                    result_excel=sheet4.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet4.min_column,max_col=sheet4.max_column)
                    for x in result_excel:
                        for y in x:
                            y.fill=greenfill
        file.save(DISK+day+'/三大法人買賣超日報3.0-'+day+'.xlsx')

    def range_Data(t,code,num):
        amount=0
        volume=0
        volume_data=[]
        midden_data={}
        for x in range(num):
            cursor.execute("select dh_difference,total_difference from {} where security_code='{}' and trade_date='{}'".format(t,code, all_day[-1-x]))
            law_data=cursor.fetchone()
            try:
                amount=amount-change(law_data[0])+change(law_data[1])
            except:
                amount=0
                break
            if t=='trading_detail':
                cursor.execute("select trade_volume,closing_price from daily_quotes where security_code='{}' and quotes_date='{}'".format(code, all_day[-1-x]))
            else:
                cursor.execute("select trade_volume,closing_price from daily_otc where security_code='{}' and quotes_date='{}'".format(code, all_day[-1-x]))
            trade_data=cursor.fetchone()
            try:
                volume_data.append(change(trade_data[0]))
                midden_data[change(trade_data[0])]=trade_data[1]
                try:
                    volume=volume+change(trade_data[0])
                except:
                    volume=0
                    break
            except:
                return ' ',' ',' '
        volume_data.sort()
        if len(volume_data)==10:
            midden=(change(midden_data.get(volume_data[4]))+change(midden_data.get(volume_data[5])))/2
        elif len(volume_data)==5:
            midden=change(midden_data.get(volume_data[2]))
        elif len(volume_data)==20:
            midden=(change(midden_data.get(volume_data[9]))+change(midden_data.get(volume_data[10])))/2
        else:
            midden=0
        try:
            rate=(amount/volume)*100
        except:
            rate=0
        return format(round(amount),','),round(midden,2),round(rate,2)
    
    cursor.execute("select trade_date from trading_detail where security_code='2330'")
    all_day=[]
    for i in cursor.fetchall():
        all_day.append(*i)
    
    for t in ['trading_detail','trading_otc']:        
        #現股
        cursor.execute("select security_code,security_name,mai_totalbuy,mai_totalsell,mai_difference,sitc_totalbuy,sitc_totalsell,sitc_difference,d_difference,dp_totalbuy,dp_totalsell,dp_difference,dh_totalbuy,dh_totalsell,dh_difference,total_difference,trade_date from {} where length(security_code)<6 and trade_date='{}'".format(t,day))
        all_data=cursor.fetchall()
        for data in all_data:
            temp=list(data)
            code=temp[0]
            total=change(temp[-2])
            dh=change(temp[-3])
            if t=='trading_detail':
                cursor.execute("select opening_price,closing_price,up_boolean,boolean,down_boolean,slope_upboolean from boolean where security_code='{}' and boolean_date='{}'".format(temp[0],day))
            else:
                cursor.execute("select opening_price,closing_price,up_boolean,boolean,down_boolean,slope_upboolean from boolean_otc where security_code='{}' and boolean_date='{}'".format(temp[0],day))
            boolean_data=cursor.fetchone()
            try:    
                opening=boolean_data[0]
                closing=boolean_data[1]
                up_boolean=boolean_data[2]
                boolean=boolean_data[3]
                down_boolean=boolean_data[4]
                slope_upbool=boolean_data[5]
            except:
                print(boolean_data,temp[0])
            try:
                temp.insert(16,round(dh/total*100,2))
            except:
                temp.insert(16,'')
            x,y,z=range_Data(t,code,5)
            x1,y1,z1=range_Data(t,code,10)
            x2,y2,z2=range_Data(t,code,20)
            temp.insert(17,x)
            temp.insert(18,y)
            temp.insert(19,z)
            temp.insert(20,x1)
            temp.insert(21,y1)
            temp.insert(22,z1)
            temp.insert(23,x2)
            temp.insert(24,y2)
            temp.insert(25,closing)
            temp.insert(26,slope_upbool)
            temp.insert(27,up_boolean)
            try:
                if float(opening)<=float(closing):
                    temp.insert(28,'紅K')
                else:
                    temp.insert(28,'黑K')
            except:
                temp.insert(28,'')
            try:
                if float(closing)>float(up_boolean):
                    temp.insert(29,'達到上布林')
                elif float(closing)<float(up_boolean) and float(closing)>float(boolean):
                    temp.insert(29,'位於中上布林區間')
                elif float(closing)<float(boolean) and float(closing)>float(down_boolean):
                    temp.insert(29,'位於中下布林區間')
                elif float(closing)<float(down_boolean):
                    temp.insert(29,'跌出下布林')
                elif float(closing)==float(up_boolean):
                    temp.insert(29,'等於上布林')
                elif float(closing)==float(boolean):
                    temp.insert(29,'等於中布林')
                elif float(closing)==float(down_boolean):
                    temp.insert(29,'等於下布林')    
            except:
                temp.insert(29,'')
            sheet.append(temp)
        #權證
        cursor.execute("select security_code,security_name,fd_totalbuy,fd_totalsell,fd_difference,d_difference,dh_totalbuy,dh_totalsell,dh_difference,total_difference,trade_date from {} where length(security_code)=6 and trade_date='{}'".format(t,day))
        all_data=cursor.fetchall()
        for data in all_data:
            temp=list(data)
            sheet1.append(temp)
    
    #標的物
    for table in[['daily_quotes','trading_detail'],['daily_otc','trading_otc']]:
        #cursor.execute("select security_code,security_name,trade_value,lbb_volume,lba_volume from {} where length(security_code)=6 and quotes_date='{}'".format(table,day))
        mysql=f"select {table[0]}.security_name,{table[0]}.trade_value,{table[0]}.lbb_volume,{table[0]}.lba_volume,{table[1]}.fd_totalbuy,{table[1]}.fd_totalsell,{table[1]}.fd_difference,{table[1]}.d_difference,{table[1]}.dh_totalbuy,{table[1]}.dh_totalsell,{table[1]}.dh_difference,{table[1]}.total_difference from {table[0]} inner join {table[1]} on {table[0]}.security_code={table[1]}.security_code and length({table[0]}.security_code)=6 and {table[0]}.quotes_date='{day}' and {table[1]}.trade_date='{day}';"
        cursor.execute(mysql)
        result_data={}
        warrant_buy=[]
        warrant_sell=[]
        warrant_cow=[]
        warrant_bear=[]
        all_data=[]
        for i in cursor.fetchall():
            if '購' in i[0]:
                warrant_buy.append(list(i))
            elif '售' in i[0]:
                warrant_sell.append(list(i))
            elif '牛' in i[0]:
                warrant_cow.append(list(i))
            elif '熊' in i[0]:
                warrant_bear.append(list(i))
        for warrant_type,sheet_num in zip([warrant_buy,warrant_sell,warrant_cow,warrant_bear],range(4)):
            subject=[]
            issue=[]
            for i in warrant_type:
                for j in ['購','售','牛','熊']:
                    if j in i[0]:
                        temp=i[0].split(j)
                        warrant_issue=temp[0][-4:-2]
                        subject_name=temp[0][0:-4]
                        subject.append(subject_name)
                        issue.append(warrant_issue)
            data_issue=list(set(issue))
            data_stock=list(set(subject))
            result=[]
            n=0
            for x in data_stock:
                n+=1
                for b in range(12):
                    #number,trade_value,lbb_volume,lba_volume,fd_totalbuy,fd_totalsell,fd_difference,d_difference,dh_totalbuy,dh_totalsell,dh_difference,total_difference
                    b=b+1    
                    locals()['total_add'+str(b)]=0
                for y in data_issue:
                    temp_data=[]
                    temp=x+y
                    # num2=0
                    for i in warrant_type:
                        if temp in i[0]:
                            temp_data.append(i)
                            # num2+=1
                    for a in range(12):
                        a=a+1    
                        locals()['add'+str(a)]=0
                    if len(temp_data)>0:
                        for j in temp_data:
                            #trade_value,lbb_volume,lba_volume,fd_totalbuy,fd_totalsell,fd_difference,d_difference,dh_totalbuy,dh_totalsell,dh_difference,total_difference
                            for a in range(11):
                                a=a+1    
                                locals()['add'+str(a)]=locals()['add'+str(a)]+change(j[a])
                                locals()['total_add'+str(a+1)]=locals()['total_add'+str(a+1)]+change(j[a])
                        locals()['total_add1']=locals()['total_add1']+int(len(temp_data))
                        result.append([stock_code(x)[0],stock_code(x)[1],y,len(temp_data),locals()['add1'],format(locals()['add2'],','),format(locals()['add3'],','),format(locals()['add4'],','),format(locals()['add5'],','),format(locals()['add6'],','),format(locals()['add7'],','),format(locals()['add8'],','),format(locals()['add9'],','),format(locals()['add10'],','),format(locals()['add11'],',')])
                cursor.execute("select closing_price,opening_price from {} where security_code='{}' and quotes_date='{}'".format(table[0],stock_code(x)[0],day))
                k_data=cursor.fetchone()
                if k_data!=None:
                    if change(k_data[0])>=change(k_data[1]):
                        k_state='紅K'
                    else:
                        k_state='黑K'
                    result.append([k_data[0],k_state,'總計',locals()['total_add1'],format(locals()['total_add2'],','),format(locals()['total_add3'],','),format(locals()['total_add4'],','),format(locals()['total_add5'],','),format(locals()['total_add6'],','),format(locals()['total_add7'],','),format(locals()['total_add8'],','),format(locals()['total_add9'],','),format(locals()['total_add10'],','),format(locals()['total_add11'],','),format(locals()['total_add12'],','),stock_code(x)[0]])
                else:
                    result.append(['','','總計',locals()['total_add1'],format(locals()['total_add2'],','),format(locals()['total_add3'],','),format(locals()['total_add4'],','),format(locals()['total_add5'],','),format(locals()['total_add6'],','),format(locals()['total_add7'],','),format(locals()['total_add8'],','),format(locals()['total_add9'],','),format(locals()['total_add10'],','),format(locals()['total_add11'],','),format(locals()['total_add12'],','),stock_code(x)[0]])   
            for i in result:
                locals()['sheet'+str(sheet_num+2)].append(i)
    
    file.save(DISK+day+'/三大法人買賣超日報3.0-'+day+'.xlsx')
    
    file=openpyxl.load_workbook(DISK+day+'/三大法人買賣超日報3.0-'+day+'.xlsx')
    sheet1=file['標的物-Top20(購)']
    sheet2=file['標的物-Top20(售)']
    sheet3=file['標的物(購)']
    sheet4=file['標的物(售)']
    sheet1.append(['股票代號','股票名稱','買進/賣出股數','買賣超股數','漲跌天數','1日漲幅','5日漲幅','自營商(避險)比例(%)','','股票代號','股票名稱','序號','買點','價格','買進股數','賣出股數'])
    sheet2.append(['股票代號','股票名稱','買進/賣出股數','買賣超股數','漲跌天數','1日漲幅','5日漲幅','自營商(避險)比例(%)','','股票代號','股票名稱','序號','買點','價格','買進股數','賣出股數'])
    
    buy_data=sheet3.values
    sell_data=sheet4.values
    buy=[]
    for i in buy_data:
        if list(i)[2]=='總計' and len(list(i)[-1])>3:
            buy.append(tuple(i))
    buy_buy20=sorted(buy,key=lambda x:change_int(x[-5]),reverse=True)[:20]
    buy_sell20=sorted(buy,key=lambda x:change_int(x[-4]),reverse=True)[:20]
    sell=[]
    for i in sell_data:
        if list(i)[2]=='總計' and len(list(i)[-1])>3:
            sell.append(tuple(i))
    sell_buy20=sorted(sell,key=lambda x:change_int(x[-5]),reverse=True)[:20]
    sell_sell20=sorted(sell,key=lambda x:change_int(x[-4]),reverse=True)[:20]
    
    final_data1=[['買賣超股數(購)-TOP20','','','(自營商買進股數)']]
    for i,x in zip([buy_buy20,buy_sell20],range(2)):
        for j in i:
            d1=increase(j[-1], 1)
            d5=increase(j[-1], 5)
            k_day=k(j[-1])
            cursor.execute("select trade_volume from daily_quotes where security_code='{}' and quotes_date='{}'".format(j[-1], day))
            volume_sum=cursor.fetchone()
            if volume_sum==None:
                cursor.execute("select trade_volume from daily_otc where security_code='{}' and quotes_date='{}'".format(j[-1], day))
                volume_sum=cursor.fetchone()
            #print(j[-3],volume_sum[0])
            diff_rate=change(j[-3])/change(volume_sum[0])*100
            if x==0:
                final_data1.append([j[-1],stock_name(j[-1]),j[-5],j[-2],k_day,d1,d5,round(diff_rate,2)])
            elif x==1:
                final_data1.append([j[-1],stock_name(j[-1]),j[-4],j[-2],k_day,d1,d5,round(diff_rate,2)])
        final_data1.append(['買賣超股數(購)-TOP20','','','(自營商賣出股數)'])
    
    final_data2=[['買賣超股數(售)-TOP20','','','(自營商買進股數)']]
    for i,x in zip([sell_buy20,sell_sell20],range(2)):
        for j in i:
            d1=increase(j[-1], 1)
            d5=increase(j[-1], 5)
            k_day=k(j[-1])
            cursor.execute("select trade_volume from daily_quotes where security_code='{}' and quotes_date='{}'".format(j[-1], day))
            volume_sum=cursor.fetchone()
            if volume_sum==None:
                cursor.execute("select trade_volume from daily_otc where security_code='{}' and quotes_date='{}'".format(j[-1], day))
                volume_sum=cursor.fetchone()
            #print(j[-3],volume_sum[0])
            diff_rate=change(j[-3])/change(volume_sum[0])*100
            if x==0:
                final_data2.append([j[-1],stock_name(j[-1]),j[-5],j[-2],k_day,d1,d5,round(diff_rate,2)])
            elif x==1:
                final_data2.append([j[-1],stock_name(j[-1]),j[-4],j[-2],k_day,d1,d5,round(diff_rate,2)])
        final_data2.append(['買賣超股數(售)-TOP20','','','(自營商賣出股數)'])
    
    for i in final_data1[:-1]:
        sheet1.append(i)
    for i in final_data2[:-1]:
        sheet2.append(i)
    
    file.save(DISK+day+'/三大法人買賣超日報3.0-'+day+'.xlsx')
    excel()
    cursor.close()
    conn.close()
    
main(day,DISK)