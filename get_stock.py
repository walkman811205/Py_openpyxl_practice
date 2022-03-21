# -*- coding: utf-8 -*-
import pymysql
import dbconfig
import openpyxl
import datetime

DISK='D:/'
d=-1
day = (datetime.datetime.now()+datetime.timedelta(days=d)).strftime("%Y%m%d")
path = DISK+day

dbconnect = dbconfig.setting
conn = pymysql.connect(**dbconnect)
cursor = conn.cursor()

def change(data):
    y=''
    for i in data:
        if i!=',':
            y=y+i
    return float(y)

def range_Data(t,code,num):
    amount=0
    volume=0
    volume_data=[]
    midden_data={}
    for x in range(num):
        cursor.execute("select dh_difference,total_difference from {} where security_code='{}' and trade_date='{}'".format(t,code, all_day[-1-x]))
        law_data=cursor.fetchone()
        try:
            amount=amount-change(law_data[0])+change(law_data[1])
        except:
            amount=0
            break
        if t=='trading_detail':
            cursor.execute("select trade_volume,closing_price from daily_quotes where security_code='{}' and quotes_date='{}'".format(code, all_day[-1-x]))
        else:
            cursor.execute("select trade_volume,closing_price from daily_otc where security_code='{}' and quotes_date='{}'".format(code, all_day[-1-x]))
        trade_data=cursor.fetchone()
        try:
            volume_data.append(change(trade_data[0]))
            midden_data[change(trade_data[0])]=trade_data[1]
            try:
                volume=volume+change(trade_data[0])
            except:
                volume=0
                break
        except:
            return ' ',' ',' '
    volume_data.sort()
    if len(volume_data)==10:
        midden=(change(midden_data.get(volume_data[4]))+change(midden_data.get(volume_data[5])))/2
    elif len(volume_data)==5:
        midden=change(midden_data.get(volume_data[2]))
    elif len(volume_data)==20:
        midden=(change(midden_data.get(volume_data[9]))+change(midden_data.get(volume_data[10])))/2
    else:
        midden=0
    try:
        rate=(amount/volume)*100
    except:
        rate=0
    return format(round(amount),','),round(midden,2),round(rate,2)   

def mark_data():
    file=openpyxl.load_workbook(DISK+day+'/現股-'+day+'.xlsx')
    greenfill=openpyxl.styles.PatternFill(fill_type='solid',fgColor='00BB00')
    redfill=openpyxl.styles.PatternFill(fill_type='solid',fgColor='FF9797')

    sheet=file['現股']
    k=sheet.iter_cols(min_row=2,max_row=sheet.max_row,min_col=29,max_col=29)
    for i in k:
        for j in i:
            if j.value=='紅K':
                result_excel=sheet.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet.min_column,max_col=sheet.max_column)
                for x in result_excel:
                    for y in x:
                        y.fill=redfill
            else:
                result_excel=sheet.iter_rows(min_row=j.row,max_row=j.row,min_col=sheet.min_column,max_col=sheet.max_column)
                for x in result_excel:
                    for y in x:
                        y.fill=greenfill
    file.save(DISK+day+'/現股-'+day+'.xlsx')



cursor.execute("select trade_date from trading_detail where security_code='2330'")
all_day=[]
for i in cursor.fetchall():
    all_day.append(*i)
    
for t in ['trading_detail','trading_otc']:        
    #現股
    file=openpyxl.load_workbook(DISK+day+'/現股-'+day+'.xlsx')
    sheet=file['現股']  
    cursor.execute("select security_code,security_name,mai_totalbuy,mai_totalsell,mai_difference,sitc_totalbuy,sitc_totalsell,sitc_difference,d_difference,dp_totalbuy,dp_totalsell,dp_difference,dh_totalbuy,dh_totalsell,dh_difference,total_difference,trade_date from {} where length(security_code)<6 and trade_date='{}'".format(t,day))
    all_data=cursor.fetchall()
    for data in all_data:
        temp=list(data)
        code=temp[0]
        total=change(temp[-2])
        dh=change(temp[-3])
        if t=='trading_detail':
            cursor.execute("select opening_price,closing_price,up_boolean,boolean,down_boolean,slope_upboolean from boolean where security_code='{}' and boolean_date='{}'".format(temp[0],day))
        else:
            cursor.execute("select opening_price,closing_price,up_boolean,boolean,down_boolean,slope_upboolean from boolean_otc where security_code='{}' and boolean_date='{}'".format(temp[0],day))
        boolean_data=cursor.fetchone()
        try:    
            opening=boolean_data[0]
            closing=boolean_data[1]
            up_boolean=boolean_data[2]
            boolean=boolean_data[3]
            down_boolean=boolean_data[4]
            slope_upbool=boolean_data[5]
        except:
            print(boolean_data,temp[0])
        try:
            temp.insert(16,round(dh/total*100,2))
        except:
            temp.insert(16,'')
            
        x,y,z=range_Data(t,code,5)
        x1,y1,z1=range_Data(t,code,10)
        x2,y2,z2=range_Data(t,code,20)
        temp.insert(17,x)
        temp.insert(18,y)
        temp.insert(19,z)
        temp.insert(20,x1)
        temp.insert(21,y1)
        temp.insert(22,z1)
        temp.insert(23,x2)
        temp.insert(24,y2)
        temp.insert(25,closing)
        temp.insert(26,slope_upbool)
        temp.insert(27,up_boolean)
        try:
            if float(opening)<=float(closing):
                temp.insert(28,'紅K')
            else:
                temp.insert(28,'黑K')
        except:
            temp.insert(28,'')
        try:
            if float(closing)>float(up_boolean):
                temp.insert(29,'達到上布林')
            elif float(closing)<float(up_boolean) and float(closing)>float(boolean):
                temp.insert(29,'位於中上布林區間')
            elif float(closing)<float(boolean) and float(closing)>float(down_boolean):
                temp.insert(29,'位於中下布林區間')
            elif float(closing)<float(down_boolean):
                temp.insert(29,'跌出下布林')
            elif float(closing)==float(up_boolean):
                temp.insert(29,'等於上布林')
            elif float(closing)==float(boolean):
                temp.insert(29,'等於中布林')
            elif float(closing)==float(down_boolean):
                temp.insert(29,'等於下布林')    
        except:
            temp.insert(29,'')
        sheet.append(temp)
    file.save(DISK+day+'/現股-'+day+'.xlsx')

mark_data()
cursor.close()
conn.close()
